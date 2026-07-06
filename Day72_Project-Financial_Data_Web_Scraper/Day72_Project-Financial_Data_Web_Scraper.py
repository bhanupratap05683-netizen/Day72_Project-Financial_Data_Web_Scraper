"""
Day 72: PROJECT - Financial Data Web Scraper (Phase 5 Capstone)
Author: Bhanu Pratap Singh

Complete pipeline:
Multi-page fetch (retry + backoff) -> BeautifulSoup parse -> validate rows
-> pandas clean/dedupe -> summary analysis -> formatted multi-sheet Excel
-> full logging + execution report.

Combines skills from Days 66-71.
"""

# ---------- IMPORTS ----------
import requests
from requests.exceptions import Timeout, ConnectionError, RequestException
from bs4 import BeautifulSoup
import pandas as pd
import logging
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# =====================================================================
# CONFIGURATION (all settings in ONE place)
# =====================================================================
CONFIG = {
    "base_url": "https://example.com/screener?page={}",  # placeholder pattern
    "pages_to_scrape": 3,
    "max_retries": 3,
    "timeout": 10,
    "delay_between_pages": 1,          # polite scraping (seconds)
    "output_file": "financial_data_report.xlsx",
    "log_file": "scraper_project.log",
    "use_sample_data": True,           # True = run offline with sample HTML
}


# =====================================================================
# SAMPLE DATA (3 "pages" so the project runs offline, always)
# =====================================================================
SAMPLE_PAGES = [
    """<html><body><table>
    <tr><th>Symbol</th><th>Company</th><th>Price</th><th>Change %</th></tr>
    <tr><td>AAPL</td><td>Apple Inc.</td><td>$195.20</td><td>1.25</td></tr>
    <tr><td>MSFT</td><td>Microsoft</td><td>$410.50</td><td>-0.80</td></tr>
    <tr><td>GOOGL</td><td>Alphabet</td><td>$152.30</td><td>2.10</td></tr>
    </table></body></html>""",

    """<html><body><table>
    <tr><th>Symbol</th><th>Company</th><th>Price</th><th>Change %</th></tr>
    <tr><td>AMZN</td><td>Amazon</td><td>$178.90</td><td>0.55</td></tr>
    <tr><td>TSLA</td><td>Tesla</td><td>bad_price</td><td>-3.20</td></tr>
    <tr><td>NVDA</td><td>NVIDIA</td><td>$885.10</td><td>4.75</td></tr>
    </table></body></html>""",

    """<html><body><table>
    <tr><th>Symbol</th><th>Company</th><th>Price</th><th>Change %</th></tr>
    <tr><td>META</td><td>Meta</td><td>$505.60</td><td>1.90</td></tr>
    <tr><td></td><td>Mystery Co</td><td>$99.99</td><td>0.10</td></tr>
    <tr><td>AAPL</td><td>Apple Inc.</td><td>$195.20</td><td>1.25</td></tr>
    </table></body></html>""",   # note: AAPL duplicate + one missing symbol
]


# =====================================================================
# LAYER 0: LOGGING SETUP (Day 71)
# =====================================================================
def setup_logging():
    """Configure logging to file and console."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(CONFIG["log_file"]),
            logging.StreamHandler(),
        ],
    )
    logging.info("=" * 60)
    logging.info("FINANCIAL DATA WEB SCRAPER - RUN STARTED")
    logging.info("=" * 60)


# =====================================================================
# LAYER 1: FETCH WITH RETRY (Day 71)
# =====================================================================
def fetch_with_retry(url):
    """Fetch a URL with exponential backoff. Raise after final failure."""
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    for attempt in range(CONFIG["max_retries"]):
        try:
            logging.info(f"Fetching {url} (attempt {attempt + 1})")
            response = requests.get(url, headers=headers,
                                    timeout=CONFIG["timeout"])
            response.raise_for_status()
            return response.text
        except (Timeout, ConnectionError, RequestException) as e:
            logging.warning(f"Attempt {attempt + 1} failed: {e}")
            if attempt < CONFIG["max_retries"] - 1:
                wait = 2 ** attempt
                logging.info(f"Retrying in {wait}s...")
                time.sleep(wait)
            else:
                logging.error(f"All retries exhausted for {url}")
                raise


# =====================================================================
# LAYER 2: MULTI-PAGE COLLECTION (Day 69)
# =====================================================================
def collect_pages():
    """Loop through pages, return list of HTML strings."""
    pages_html = []
    for page_num in range(1, CONFIG["pages_to_scrape"] + 1):
        if CONFIG["use_sample_data"]:
            logging.info(f"Loading sample page {page_num}")
            pages_html.append(SAMPLE_PAGES[page_num - 1])
        else:
            url = CONFIG["base_url"].format(page_num)
            try:
                pages_html.append(fetch_with_retry(url))
            except RequestException:
                logging.error(f"Skipping page {page_num} after failures")
                continue  # partial success: keep other pages
        # Polite pause between pages
        if page_num < CONFIG["pages_to_scrape"]:
            time.sleep(CONFIG["delay_between_pages"])
    logging.info(f"Collected {len(pages_html)} pages")
    return pages_html


# =====================================================================
# LAYER 3: PARSE TABLE (Days 67-68)
# =====================================================================
def parse_table(html, page_label):
    """Extract raw rows from the first table. Return list of cell lists."""
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if table is None:
        logging.error(f"{page_label}: no table found")
        return []
    raw_rows = []
    for tr in table.find_all("tr")[1:]:            # skip header row
        cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
        raw_rows.append(cells)
    logging.info(f"{page_label}: extracted {len(raw_rows)} raw rows")
    return raw_rows


# =====================================================================
# LAYER 4: VALIDATE ROWS (Day 71 - partial success pattern)
# =====================================================================
def safe_float(value):
    """Convert '$195.20' or '1.25' to float. Return None if invalid."""
    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def validate_rows(raw_rows):
    """Keep only complete, numeric rows. Log and count skipped rows."""
    valid, skipped = [], 0
    for i, row in enumerate(raw_rows, start=1):
        if len(row) < 4:
            logging.warning(f"Row {i}: wrong column count {row}. Skipped.")
            skipped += 1
            continue
        symbol, company = row[0], row[1]
        price = safe_float(row[2])
        change = safe_float(row[3])
        if not symbol or price is None or change is None:
            logging.warning(f"Row {i}: invalid data {row}. Skipped.")
            skipped += 1
            continue
        valid.append([symbol, company, price, change])
    logging.info(f"Validation: {len(valid)} valid, {skipped} skipped")
    return valid, skipped


# =====================================================================
# LAYER 5: CLEAN WITH PANDAS (Day 70)
# =====================================================================
def build_dataframe(valid_rows):
    """Create DataFrame, remove duplicate symbols, sort by price."""
    df = pd.DataFrame(valid_rows,
                      columns=["Symbol", "Company", "Price", "Change %"])
    before = len(df)
    df = df.drop_duplicates(subset="Symbol", keep="first")
    dupes = before - len(df)
    if dupes:
        logging.info(f"Removed {dupes} duplicate symbol(s)")
    df = df.sort_values("Price", ascending=False).reset_index(drop=True)
    return df


# =====================================================================
# LAYER 6: ANALYSIS / SUMMARY (the analyst part)
# =====================================================================
def build_summary(df):
    """Compute summary statistics as a small DataFrame."""
    summary = {
        "Metric": ["Total Stocks", "Average Price", "Highest Price",
                   "Lowest Price", "Gainers", "Losers",
                   "Avg Change %", "Report Generated"],
        "Value": [
            len(df),
            round(df["Price"].mean(), 2),
            df["Price"].max(),
            df["Price"].min(),
            int((df["Change %"] > 0).sum()),
            int((df["Change %"] < 0).sum()),
            round(df["Change %"].mean(), 2),
            time.strftime("%Y-%m-%d %H:%M:%S"),
        ],
    }
    return pd.DataFrame(summary)


def build_top_movers(df, n=3):
    """Top N stocks by absolute % change."""
    return (df.reindex(df["Change %"].abs()
                       .sort_values(ascending=False).index)
              .head(n).reset_index(drop=True))


# =====================================================================
# LAYER 7: EXPORT MULTI-SHEET EXCEL + FORMATTING (Days 70, 13, 18)
# =====================================================================
def export_report(df, summary_df, movers_df):
    """Write 3-sheet Excel workbook, then apply professional styling."""
    file = CONFIG["output_file"]

    with pd.ExcelWriter(file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Market Data", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        movers_df.to_excel(writer, sheet_name="Top Movers", index=False)
    logging.info(f"Workbook written: {file}")

    # ---- Formatting pass ----
    wb = load_workbook(file)
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="1F4E2C")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        for cell in ws[1]:                        # header row
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        for row in ws.iter_rows(min_row=2):       # data rows
            for cell in row:
                cell.border = border
        for col in ws.columns:                    # auto column width
            width = max(len(str(c.value)) for c in col if c.value) + 4
            ws.column_dimensions[col[0].column_letter].width = width

    wb.save(file)
    logging.info("Formatting applied to all sheets")
    return file


# =====================================================================
# MAIN ORCHESTRATOR
# =====================================================================
def main():
    start = time.time()
    print("\n" + "=" * 60)
    print("DAY 72 PROJECT: FINANCIAL DATA WEB SCRAPER")
    print("=" * 60)

    try:
        # 1. Collect pages
        pages = collect_pages()
        if not pages:
            print("❌ No pages could be fetched. See log.")
            return

        # 2. Parse all pages
        raw_rows = []
        for i, html in enumerate(pages, start=1):
            raw_rows.extend(parse_table(html, f"Page {i}"))
        if not raw_rows:
            print("❌ No data rows found on any page. See log.")
            return

        # 3. Validate
        valid_rows, skipped = validate_rows(raw_rows)
        if not valid_rows:
            print("❌ All rows failed validation. See log.")
            return

        # 4. Clean
        df = build_dataframe(valid_rows)

        # 5. Analyze
        summary_df = build_summary(df)
        movers_df = build_top_movers(df)

        # 6. Export
        file = export_report(df, summary_df, movers_df)

        # 7. Execution report
        elapsed = time.time() - start
        print(f"\n✅ PROJECT RUN COMPLETE  ({elapsed:.1f}s)")
        print(f"   Pages scraped : {len(pages)}")
        print(f"   Raw rows      : {len(raw_rows)}")
        print(f"   Valid rows    : {len(df)} (skipped {skipped}, "
              f"deduped {len(valid_rows) - len(df)})")
        print(f"   Output file   : {file}")
        print(f"   Log file      : {CONFIG['log_file']}\n")
        print("--- Market Data Preview ---")
        print(df.to_string(index=False))
        print("\n--- Summary ---")
        print(summary_df.to_string(index=False))
        logging.info("RUN COMPLETED SUCCESSFULLY")

    except Exception as e:
        logging.exception(f"UNEXPECTED FAILURE: {e}")
        print(f"\n❌ Unexpected error: {e}. Full traceback in log.")


if __name__ == "__main__":
    setup_logging()
    main()